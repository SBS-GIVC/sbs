#!/usr/bin/env python3
"""
SBS Data Processor
Converts official CHI SBS Excel files to JSON format for integration
"""

import json
import os
from openpyxl import load_workbook
from datetime import datetime

OUTPUT_DIR = "processed"

def ensure_output_dir():
    """Ensure output directory exists"""
    if not os.path.exists(OUTPUT_DIR):
        os.makedirs(OUTPUT_DIR)

def process_sbs_code_list():
    """Process the main SBS V3 Electronic Code List"""
    print("Processing SBS V3 Electronic Code List...")
    
    try:
        wb = load_workbook("SBS_V3_Electronic_Code_List.xlsx", data_only=True)
        
        all_codes = []
        metadata = {
            "source": "CHI - Council of Health Insurance",
            "version": "SBS V3.1",
            "processed_at": datetime.now().isoformat(),
            "sheets": []
        }
        
        for sheet_name in wb.sheetnames:
            print(f"  Processing sheet: {sheet_name}")
            ws = wb[sheet_name]
            
            # Get header row (first row)
            headers = []
            for col in range(1, ws.max_column + 1):
                cell_value = ws.cell(row=1, column=col).value
                headers.append(str(cell_value) if cell_value else f"column_{col}")
            
            # Process data rows
            sheet_codes = []
            for row in range(2, ws.max_row + 1):
                row_data = {}
                for col, header in enumerate(headers, 1):
                    cell_value = ws.cell(row=row, column=col).value
                    if cell_value is not None:
                        row_data[header] = str(cell_value).strip() if isinstance(cell_value, str) else cell_value
                
                # Only include rows that have meaningful data
                if row_data and any(v for v in row_data.values() if v):
                    # Try to identify the SBS code column
                    code = None
                    for key in row_data:
                        key_lower = key.lower()
                        if 'sbs' in key_lower and 'code' in key_lower:
                            code = row_data[key]
                            break
                        elif 'code' in key_lower and code is None:
                            code = row_data[key]
                    
                    if code:
                        sheet_codes.append(row_data)
            
            if sheet_codes:
                metadata["sheets"].append({
                    "name": sheet_name,
                    "count": len(sheet_codes),
                    "headers": headers
                })
                all_codes.extend(sheet_codes)
            
            print(f"    Found {len(sheet_codes)} codes")
        
        # Save raw data
        with open(f"{OUTPUT_DIR}/sbs_codes_raw.json", "w", encoding="utf-8") as f:
            json.dump({
                "metadata": metadata,
                "codes": all_codes
            }, f, ensure_ascii=False, indent=2)
        
        print(f"  Total codes extracted: {len(all_codes)}")
        return all_codes
        
    except Exception as e:
        print(f"  Error processing SBS code list: {e}")
        return []

def process_snomed_mapping():
    """Process SBS to SNOMED mapping"""
    print("Processing SBS to SNOMED Map...")
    
    try:
        wb = load_workbook("SBS_V3_to_SNOMED_Map.xlsx", data_only=True)
        
        mappings = []
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            
            headers = []
            for col in range(1, ws.max_column + 1):
                cell_value = ws.cell(row=1, column=col).value
                headers.append(str(cell_value) if cell_value else f"column_{col}")
            
            for row in range(2, ws.max_row + 1):
                row_data = {}
                for col, header in enumerate(headers, 1):
                    cell_value = ws.cell(row=row, column=col).value
                    if cell_value is not None:
                        row_data[header] = str(cell_value).strip() if isinstance(cell_value, str) else cell_value
                
                if row_data and any(v for v in row_data.values() if v):
                    mappings.append(row_data)
        
        with open(f"{OUTPUT_DIR}/sbs_snomed_map.json", "w", encoding="utf-8") as f:
            json.dump({
                "source": "CHI SBS V3 to SNOMED Map",
                "count": len(mappings),
                "mappings": mappings
            }, f, ensure_ascii=False, indent=2)
        
        print(f"  Total SNOMED mappings: {len(mappings)}")
        return mappings
        
    except Exception as e:
        print(f"  Error processing SNOMED map: {e}")
        return []

def process_achi_mapping():
    """Process SBS to ACHI mapping"""
    print("Processing SBS to ACHI Map...")
    
    try:
        wb = load_workbook("SBS_V3_to_ACHI_Map.xlsx", data_only=True)
        
        mappings = []
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            
            headers = []
            for col in range(1, ws.max_column + 1):
                cell_value = ws.cell(row=1, column=col).value
                headers.append(str(cell_value) if cell_value else f"column_{col}")
            
            for row in range(2, ws.max_row + 1):
                row_data = {}
                for col, header in enumerate(headers, 1):
                    cell_value = ws.cell(row=row, column=col).value
                    if cell_value is not None:
                        row_data[header] = str(cell_value).strip() if isinstance(cell_value, str) else cell_value
                
                if row_data and any(v for v in row_data.values() if v):
                    mappings.append(row_data)
        
        with open(f"{OUTPUT_DIR}/sbs_achi_map.json", "w", encoding="utf-8") as f:
            json.dump({
                "source": "CHI SBS V3 to ACHI 10th Edition Map",
                "count": len(mappings),
                "mappings": mappings
            }, f, ensure_ascii=False, indent=2)
        
        print(f"  Total ACHI mappings: {len(mappings)}")
        return mappings
        
    except Exception as e:
        print(f"  Error processing ACHI map: {e}")
        return []

def process_dental_pricelist():
    """Process Dental Services Price List"""
    print("Processing Dental Services Pricelist...")
    
    try:
        wb = load_workbook("Dental_Services_Pricelist.xlsx", data_only=True)
        
        services = []
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            
            headers = []
            for col in range(1, ws.max_column + 1):
                cell_value = ws.cell(row=1, column=col).value
                headers.append(str(cell_value) if cell_value else f"column_{col}")
            
            for row in range(2, ws.max_row + 1):
                row_data = {}
                for col, header in enumerate(headers, 1):
                    cell_value = ws.cell(row=row, column=col).value
                    if cell_value is not None:
                        row_data[header] = str(cell_value).strip() if isinstance(cell_value, str) else cell_value
                
                if row_data and any(v for v in row_data.values() if v):
                    services.append(row_data)
        
        with open(f"{OUTPUT_DIR}/dental_pricelist.json", "w", encoding="utf-8") as f:
            json.dump({
                "source": "CHI Dental Services Pricelist (Government Sector)",
                "count": len(services),
                "services": services
            }, f, ensure_ascii=False, indent=2)
        
        print(f"  Total dental services: {len(services)}")
        return services
        
    except Exception as e:
        print(f"  Error processing dental pricelist: {e}")
        return []

def create_normalized_db():
    """Create a normalized database structure for the application"""
    print("\nCreating normalized database structure...")
    
    ensure_output_dir()
    
    # Load the raw SBS codes
    try:
        with open(f"{OUTPUT_DIR}/sbs_codes_raw.json", "r", encoding="utf-8") as f:
            data = json.load(f)
            raw_codes = data.get("codes", [])
    except:
        raw_codes = []
    
    # Create a simplified structure for the app
    sbs_catalogue = []
    
    for code_data in raw_codes:
        # Extract relevant fields
        entry = {
            "sbs_id": None,
            "sbs_code": None,
            "description_en": None,
            "description_ar": None,
            "category": None,
            "subcategory": None,
            "unit": None,
            "is_active": True
        }
        
        # Map fields based on common column names
        for key, value in code_data.items():
            key_lower = key.lower()
            
            if 'sbs' in key_lower and 'code' in key_lower:
                entry["sbs_code"] = value
                entry["sbs_id"] = value
            elif 'description' in key_lower and 'english' in key_lower:
                entry["description_en"] = value
            elif 'description' in key_lower and 'arabic' in key_lower:
                entry["description_ar"] = value
            elif 'description' in key_lower and entry["description_en"] is None:
                entry["description_en"] = value
            elif 'category' in key_lower and entry["category"] is None:
                entry["category"] = value
            elif 'block' in key_lower:
                entry["subcategory"] = value
            elif 'unit' in key_lower:
                entry["unit"] = value
        
        if entry["sbs_code"]:
            sbs_catalogue.append(entry)
    
    # Save the catalogue
    with open(f"{OUTPUT_DIR}/sbs_catalogue.json", "w", encoding="utf-8") as f:
        json.dump({
            "version": "3.1",
            "source": "CHI Official SBS V3",
            "generated_at": datetime.now().isoformat(),
            "total_codes": len(sbs_catalogue),
            "catalogue": sbs_catalogue
        }, f, ensure_ascii=False, indent=2)
    
    print(f"  Created catalogue with {len(sbs_catalogue)} entries")
    
    # Create category index
    categories = {}
    for entry in sbs_catalogue:
        cat = entry.get("category", "Unknown")
        if cat not in categories:
            categories[cat] = []
        categories[cat].append(entry["sbs_code"])
    
    with open(f"{OUTPUT_DIR}/sbs_categories.json", "w", encoding="utf-8") as f:
        json.dump(categories, f, ensure_ascii=False, indent=2)
    
    print(f"  Created {len(categories)} category indexes")

def main():
    """Main processing function"""
    print("=" * 60)
    print("SBS Data Processor")
    print("Processing official CHI Saudi Billing System files")
    print("=" * 60)
    print()
    
    ensure_output_dir()
    
    # Process all files
    process_sbs_code_list()
    process_snomed_mapping()
    process_achi_mapping()
    process_dental_pricelist()
    
    # Create normalized structure
    create_normalized_db()
    
    print()
    print("=" * 60)
    print("Processing complete!")
    print(f"Output files saved to: {OUTPUT_DIR}/")
    print("=" * 60)

if __name__ == "__main__":
    main()
